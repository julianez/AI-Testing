import torch
import logging
import numpy as np
import pandas as pd
import torch.nn.functional as F
from transformers import LlamaForCausalLM, LlamaTokenizer, BitsAndBytesConfig
from openpyxl import load_workbook

def getTokenIds(concept,question):
    # Tokenize inputs
    tokenizer.pad_token = tokenizer.eos_token

    concept_ids = tokenizer(concept, return_tensors='pt')['input_ids'].to(device)
    question_ids = tokenizer(question, return_tensors='pt')['input_ids'].to(device)

    return concept_ids, question_ids

def padEmbeddings(embedding1,embedding2):
    logging.debug("Shape (Before) of embedding1_cpu: %s", embedding1.shape)
    logging.debug("Shape (Before) of embedding2_cpu: %s", embedding2.shape)

    if embedding2.size(1) > embedding1.size(1):
        # Calculate the difference in size along the second dimension
        diff = embedding2.size(1) - embedding1.size(1)
        # Create the padding tuple
        padding = (0, 0, 0, diff)
        # Pad the smaller tensor
        embedding1 = F.pad(embedding1, padding)
    else:
        # Calculate the difference in size along the second dimension
        diff = embedding1.size(1) - embedding2.size(1)
        # Create the padding tuple
        padding = (0, 0, 0, diff)
        # Pad the smaller tensor
        embedding2 = F.pad(embedding2, padding)


    logging.debug("Shape (After) of embedding1_cpu: %s", embedding1.shape)
    logging.debug("Shape (After) of embedding2_cpu: %s", embedding2.shape)

    return embedding1,embedding2

def combineEmbeddingMethod1(embedding1,embedding2,weight1,weight2):
    #Weighted Concatenation with Numpy
    embedding1_cpu = embedding1.cpu().detach()
    embedding2_cpu = embedding2.cpu().detach()

    weighted_embedding_np = np.concatenate([embedding1_cpu * weight1 , embedding2_cpu * weight2])
    weighted_embedding = torch.from_numpy(weighted_embedding_np).to(device)
    
    return weighted_embedding

def combineEmbeddingMethod2(embedding1,embedding2,weight1,weight2):
    #Weighted Concatenation with PyTorch
    concatenated_embedding = torch.cat([embedding1*weight1, embedding2*weight2], dim=1)
    return concatenated_embedding

def combineEmbeddingMethod3(embedding1,embedding2):
    #Element-wise Multiplication
    element_wise_embedding = embedding1 * embedding2
    return element_wise_embedding

def combineEmbeddingMethod4(embedding1,embedding2,weight1,weight2):
    # 2. Element-wise addition:
    combined_embedding = embedding1*weight1 + embedding2*weight2

    return combined_embedding

def combineEmbeddingMethod5(embedding1,embedding2,weight1,weight2):
    # 4. Average:
    combined_embedding = (embedding1*weight1 + embedding2*weight2) / 2

    return combined_embedding

LOGGING=logging
logging.basicConfig(level=logging.ERROR, format='%(asctime)s - %(levelname)s - %(message)s')

device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
logging.debug("Device: %s", device)

# Load the Llama-2-7b model and tokenizer
model_path = "meta-llama/Llama-2-7b-hf"
#model_path = 'openlm-research/open_llama_7b_v2'
#model_path = "openlm-research/open_llama_3b_v2"
#model_path = "TinyLlama/TinyLlama-1.1B-Chat-v1.0"
#model_path = "TinyLlama/TinyLlama-1.1B-intermediate-step-1431k-3T"

bnb_config = BitsAndBytesConfig(
    load_in_4bit=True,
    bnb_4bit_use_double_quant=True,
    bnb_4bit_quant_type="nf4",
    bnb_4bit_compute_dtype=torch.bfloat16
)

tokenizer = LlamaTokenizer.from_pretrained(model_path,legacy=False)
model = LlamaForCausalLM.from_pretrained(model_path,
                                         quantization_config=bnb_config,
                                         torch_dtype=torch.float16, 
                                         device_map='auto')

# Inputs
concept = "Biodiversity "
question = "The capital of Colombia is "

concept_ids, question_ids = getTokenIds(concept,question)

# Access model layers
embeddings = model.get_input_embeddings()
embeddings = embeddings.to(device)

# Get embedded representation of the ids
embedded_concept = embeddings(concept_ids)
embedded_question = embeddings(question_ids)

embedded_concept,embedded_question = padEmbeddings(embedded_concept,embedded_question)

combinedEmbedding = []

#0
combinedEmbedding.append(embedded_question)

#1
combinedEmbedding.append(combineEmbeddingMethod1(embedded_concept,embedded_question,1,1))

#2 - 6
combinedEmbedding.append(combineEmbeddingMethod2(embedded_concept,embedded_question,1,1))
combinedEmbedding.append(combineEmbeddingMethod2(embedded_concept,embedded_question,2,1))
combinedEmbedding.append(combineEmbeddingMethod2(embedded_concept,embedded_question,1,2))
combinedEmbedding.append(combineEmbeddingMethod2(embedded_concept,embedded_question,1,3))
combinedEmbedding.append(combineEmbeddingMethod2(embedded_concept,embedded_question,3,1))

#7
combinedEmbedding.append(combineEmbeddingMethod3(embedded_concept,embedded_question))

#8 - 12
combinedEmbedding.append(combineEmbeddingMethod4(embedded_concept,embedded_question,1,1))
combinedEmbedding.append(combineEmbeddingMethod4(embedded_concept,embedded_question,2,1))
combinedEmbedding.append(combineEmbeddingMethod4(embedded_concept,embedded_question,1,2))
combinedEmbedding.append(combineEmbeddingMethod4(embedded_concept,embedded_question,1,3))
combinedEmbedding.append(combineEmbeddingMethod4(embedded_concept,embedded_question,3,1))

#13 - 17
combinedEmbedding.append(combineEmbeddingMethod5(embedded_concept,embedded_question,1,1))
combinedEmbedding.append(combineEmbeddingMethod5(embedded_concept,embedded_question,2,1))
combinedEmbedding.append(combineEmbeddingMethod5(embedded_concept,embedded_question,1,2))
combinedEmbedding.append(combineEmbeddingMethod5(embedded_concept,embedded_question,1,3))
combinedEmbedding.append(combineEmbeddingMethod5(embedded_concept,embedded_question,3,1))

output = []

for index,combined in enumerate(combinedEmbedding):
    # https://huggingface.co/docs/transformers/v4.18.0/en/main_classes/text_generation
    logging.debug("Combined Type: %s", type(combined))
    logging.debug("Combined Embedding: %s ",combined)
    with torch.no_grad():
        generated_ids = model.generate(
            #input_ids = input_ids,
            inputs_embeds = combined,
            num_return_sequences=1,
            min_length=1,
            max_length=250,
            do_sample=True,
            top_p=0.9,
            temperature=0.3,
            num_beams = 6, 
            length_penalty= 1.5,
            no_repeat_ngram_size=1,
        )
    generated_text = tokenizer.decode(generated_ids[0], skip_special_tokens=True)
    print(f"\nCombination {index}: {generated_text}")
    output.append((concept,
                   question,
                   index,
                   generated_text))

# Convert list to numpy array
output_np = np.array(output)
# Convert numpy array to DataFrame
df = pd.DataFrame(output_np)

# Try to load existing workbook
try:
    book = load_workbook('combine-embeddings.xlsx')
    # Create a pandas Excel writer using openpyxl as the engine
    writer = pd.ExcelWriter('combine-embeddings.xlsx', engine='openpyxl')
    # Check if 'Sheet1' exists in the workbook, if not create it
    if 'Sheet1' not in writer.book.sheetnames:
        writer.book.create_sheet('Sheet1')
    startrow = writer.sheets['Sheet1'].max_row
except FileNotFoundError:
    # File does not exist yet, we will create it in the next step
    writer = pd.ExcelWriter('combine-embeddings.xlsx', engine='openpyxl')
    startrow = 0

# Write DataFrame to the excel file, starting at the last row of the existing data
df.to_excel(writer, startrow=startrow, index=False, header=False)

# Save the changes
writer.book.save('combine-embeddings.xlsx')